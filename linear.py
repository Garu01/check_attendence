import tkinter
from tkinter import *
from tkinter.ttk import *
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from PIL import ImageTk,Image
import datetime
import time

#create Gui
window = Tk()
window.title("Check attendance")
window.geometry("600x400")
photo = PhotoImage(file="vgu_logo")
window.iconphoto(False,photo)

#connect to excel file
wb = Workbook()
wb = load_workbook("hello.xlsx")

#print from excel file function
def query():

    sheet_name = combo.get()
    if sheet_name in ws:

        data = pd.read_excel('hello.xlsx',sheet_name=sheet_name)

        get_name = data['Name'].dropna()
        set_name = get_name.value_counts().sort_index().index.tolist()
        set_name = list(map(str, set_name))

        get_check = data['Check'].dropna()
        set_check = get_check.values.tolist()
        set_check = list(map(str, set_check))

        out = ""
        for i in range(0,len(set_name)) :
             out +=   "{:<20}".format(set_name[i]) +  "{:>10}".format(set_check[i])+"\n"

        a.configure(text=out)

#====================================================================

#label
a = tkinter.Label(window,text="  ",fg="blue",font=("Ariel",10),justify=LEFT)
a.place(x= 250,y=150)

#create bar chart
#This bar chart perform a number of student attending class or not
def graph():
    sheet_name = str(combo.get())
    if sheet_name in wb.sheetnames:
        #choose read the sheet that was chosen by user
        data = pd.read_excel('hello.xlsx',sheet_name=sheet_name)

        #get value from check column
        get_check_value = data['Check'].dropna()

        plt.figure(figsize=(15, 5))
        check_value= get_check_value.value_counts().sort_index().index.tolist()
        check_value = list(map(str, check_value))
        plt.bar(check_value, get_check_value.value_counts().sort_index())
        plt.grid()
        plt.bar(check_value, get_check_value.value_counts().sort_index(), color='g')
        # plt.legend('Number of Student')
        date = datetime.date.today()
        plt.xlabel(date)
        plt.ylabel('Number of student')
        plt.title('Check Attendance result')
        plt.show()
#=================================================================

#set color
canvas = Canvas(window,bg="red",width=200,height=600)
canvas.pack(anchor=NW)
canvas_1 = Canvas(window,bg="orange",width=400,height=100)
canvas_1.place(x=202,y=0)
name = tkinter.Label(window,bg="orange",text="Check attendance",font=("Ariel",30))
name.place(x=250,y=20)
#add vgu logo
image = Image.open("vgu_logo")
image = image.resize((160,110))
image_gui = ImageTk.PhotoImage(image)
image_label = Label(image=image_gui)
image_label.place(x=20,y=10)

#create label date
date = tkinter.Label(window,text="Date: ",bg="red")
date.place(x=5,y=150)
#combobox
combo = Combobox(window)

#read file name from exel
ws  = wb.sheetnames
combo["value"] = ws
combo.place(x=35,y=150,width=100,height=30)
#================================================================


#create a show button
show_combo = tkinter.Button(window,text="Show record",command=query)
show_combo.place(x=35,y=200,width=100,height=30)

#create graph button
graph_button = tkinter.Button(window,text="Show graph",command=graph)
graph_button.place(x=35,y=230,width=100,height=30)

#exit button
exit_button = tkinter .Button (text = " Exit ", command= window.destroy)
exit_button.place (x = 5 ,y = 340 ,width =100 ,height = 50)


window.mainloop()