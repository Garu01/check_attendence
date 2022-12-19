from keras.models import load_model
from PIL import Image, ImageOps
import numpy as np
import cv2
from openpyxl import load_workbook
import datetime
import time

# Load the model
model = load_model('keras_model.h5')

def image_capture():
    cam=cv2.VideoCapture(0)
    ret,frame = cam.read()
    cv2.imwrite("test.png",frame)

def image_detector():
    # Create the array of the right shape to feed into the keras model
    # The 'length' or number of images you can put into the array is
    # determined by the first position in the shape tuple, in this case 1.
    data = np.ndarray(shape=(1, 224, 224, 3), dtype=np.float32)
    # Replace this with the path to your image
    image = Image.open('test.png')
    # resize the image to a 224x224 with the same strategy as in TM2:
    # resizing the image to be at least 224x224 and then cropping from the center
    size = (224, 224)
    image = ImageOps.fit(image, size, Image.ANTIALIAS)

    # turn the image into a numpy array
    image_array = np.asarray(image)
    # Normalize the image
    normalized_image_array = (image_array.astype(np.float32) / 127.0) - 1
    # Load the image into the array
    data[0] = normalized_image_array

    # run the inference
    prediction = model.predict(data)
    print(prediction)

    # Find the maximum confidence and its index from 2D array
    # get the 1D array
    # prediction dang la [[... ... ...]]
    output = prediction[0]
    # assign default value for max confidence
    global max_index
    max_index = -1
   # max_confidence = output[0]
    max_confidence = 0

    # find the maximum confidence and its index
    for i in range(0, len(output)):
        if output[i] > 0.9 :
            max_confidence = output[i]
            max_index = i

    if max_index != -1:
        print(max_index, max_confidence)
        file = open("labels.txt", encoding="utf-8")
        data = file.read().split("\n")
        print("AI result : ", data[max_index])
    else :
        print(" Can not detect")



file = open("labels.txt", encoding="utf-8")
data = file.read().split("\n")
global max_index


arr = [0,0,0,0,0,0]
#connect to xlsx file, setup name each sheet
wb = load_workbook("hello.xlsx")
date = datetime.date.today()
wb.create_sheet(str(date))
sheets = wb.sheetnames
sheet_today = wb[str(date)]
sheet_today.cell(row=1,column=1).value = 'Name'
sheet_today.cell(row=1,column=2).value = 'Check'
for i in range(0,3):
    sheet_today.cell(row=i+2, column=1).value = data[i]
    sheet_today.cell(row=i+2, column=2).value = 0
wb.save('hello.xlsx')

while True:
    #capture image each 3 seconds
    time.sleep(3)
    image_capture()
    image_detector()
    if(max_index != -1) :
        arr[max_index] +=1
        if arr[max_index] >= 3 :
            sheet_today.cell(row=max_index+2, column=2).value = 1
            wb.save("hello.xlsx")
            print(data[max_index] + " has been attended! ")
workbook.close()






